import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
import io


# Fonction de traitement pour le fichier Excel
def process_excel(file):
    try:
        # Lire le fichier Excel sans header
        xls = pd.read_excel(file, header=None)
                
        # Trouver l'index de la ligne d'en-tête
        header_index = next(
            (idx for idx, row in xls.iterrows() if all(col in row.values for col in ['Code', 'Nom', 'Prénom'])),
            None
        )
        
        if header_index is None:
            st.error("Les colonnes 'Code', 'Nom', 'Prénom' sont introuvables dans le fichier.")
            return None, None
        
        # Redéfinir les en-têtes et supprimer les lignes précédentes
        xls.columns = xls.iloc[header_index]
        xls = xls.iloc[header_index + 1:].reset_index(drop=True)
        
        # Vérification si le fichier est vide après nettoyage
        if xls.empty:
            st.error("Aucune donnée valide après le traitement des lignes.")
            return None, None
        
        # Vérification des colonnes nécessaires (double vérification)
        required_columns = ['Nom', 'Prénom', 'Code']
        missing = [col for col in required_columns if col not in xls.columns]
        if missing:
            st.error(f"Colonnes manquantes après traitement : {', '.join(missing)}")
            return None, None
        
        # Nettoyage des données
        liste = xls.dropna(subset=['Nom', 'Prénom', 'Code'])
        liste['Name'] = liste['Code'].astype(str) + ' ' + liste['Nom'] + ' ' + liste['Prénom']
        liste = liste[['Code', 'Name']].drop_duplicates()
        return xls, liste
    
    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Excel : {str(e)}")
        st.info("Assurez-vous que le fichier est bien formaté et contient les colonnes requises.")
        return None, 

# Fonction de traitement pour le fichier CSV
def process_csv2excel(xls_file, csv_file):
    try:
        csv = pd.read_csv(csv_file, delimiter=';', encoding='utf-8')

        # Nettoyer les données : supprimer les lignes où 'A:Code' == 'NONE'
        anomalies = csv[csv['A:Code'] == 'NONE'].copy()
        csv_clean = csv[csv['A:Code'] != 'NONE'].copy()

        # Vérifier si le fichier nettoyé est vide
        if csv_clean.empty:
            print("Aucune donnée valide après le nettoyage !")

        # Construire le dictionnaire Notes
        notes_etudiants = {row['A:Code'].strip().upper(): row['Note'] for _, row in csv_clean.iterrows()}
        notes_etudiants = {int(cle): valeur for cle, valeur in notes_etudiants.items()}

        wb = load_workbook(filename=xls_file)
        sheet = wb.active # ou workbook['Nom_de_la_feuille']
        cell_row = None
        cell_column = None

        # Parcourir toutes les lignes et colonnes de la feuille de calcul
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 'Note':  # Vérifier si la valeur de la cellule est 'Note'
                    cell_row = cell.row  # Ligne de la cellule
                    cell_column = cell.column  # Colonne de la cellule
                    #st.success(
                    #    f"L'en-tête 'Note' a été trouvée à la ligne {cell_row}, colonne {cell_column} ({cell.coordinate}).")
                    break  # Sortir de la boucle interne après avoir trouvé la valeur
            if cell_row is not None and cell_column is not None:
                break  # Sortir de la boucle externe après avoir trouvé la valeur

        # Si la valeur n'a pas été trouvée
        if cell_row is None and cell_column is None:
            st.warning("La chaîne 'Note' n'a pas été trouvée dans le fichier.")

        # Parcourir les lignes du fichier Excel
        for row in sheet.iter_rows(min_row=cell_row, max_col=cell_column, values_only=False):  # Ignorer l'en-tête
            code_etudiant_cell = row[0]  # Colonne des codes d'étudiants (première colonne)
            note_cell = row[cell_column-1]  # Colonne des notes (deuxième colonne)

            # Vérifier si le code étudiant existe dans le dictionnaire
            if code_etudiant_cell.value in notes_etudiants:
                note_cell.value = notes_etudiants[code_etudiant_cell.value]

        # Enregistrer les modifications dans le fichier Excel
        wb.save(xls_file)
        wb.close()
        return xls_file, len(anomalies)

    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Excel : {str(e)}")
        st.error("Assurez-vous que le fichier est bien formaté et contient les colonnes requises.")
        return None, None



# ----------------- Interface utilisateur -----------------
st.title("🛠️ Outils pour AMC")

# Sidebar pour les sections
section = st.sidebar.radio(" ", ["Étudiants", "Notes", "Statistiques"])

if section == "Étudiants":
    st.header("👨‍🎓 Liste des étudiants")
    st.info(
        """
        Télécharger le fichier Excel de l'administration.
        """
    )
    
    uploaded_excel_file = st.file_uploader(
        " ",
        type="xlsx", 
        key="excel_uploader"
    )
    
    if uploaded_excel_file is not None:
        with st.spinner("Traitement automatique du fichier Excel en cours..."):
            xls, liste = process_excel(uploaded_excel_file)
            
            if xls is not None:
                st.success(f"Lecture du fichier Excel réussie ! {len(xls)} étudiants trouvés.")  
                st.write("🔎 Aperçu de la base de données des étudiants:")
                st.write(xls.head(10))              
                st.write("🔎 Aperçu de la liste des étudiants à fournir à AMC:")
                st.write(liste.head(10))
                st.success(f"La liste contient {len(xls)} étudiants.")
                # Générer le fichier CSV
                csv_data = liste.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📥 Télécharger la liste des étudiants au format CSV",
                    data=csv_data,
                    file_name="liste_etudiants.csv",
                    mime="text/csv"
                )

elif section == "Notes":
    st.header("✍️ Traitement des notes")
    st.info(
        """
        Télécharger le fichier Excel de l'administration.
        """
    )
    xls_file = st.file_uploader(
        " ",
        type="xlsx", 
        key="excel_uploader2"
    )

    st.info(
        """
        Télécharger le fichier CSV des notes calculées par AMC
        """
    )
    csv_file = st.file_uploader(
        "",
        type="csv", 
        key="csv_uploader"
    )
    if xls_file is not None and csv_file is not None:
        with st.spinner("Traitement automatique du fichier Excel en cours..."):
            # Lire le contenu du fichier CSV en mémoire
            csv_content = io.BytesIO(csv_file.getvalue())

            # Lire le contenu du fichier Excel en mémoire
            xls_content = io.BytesIO(xls_file.getvalue())

            # Traitez les fichiers avec process_csv2excel
            processed_xls, nb_none = process_csv2excel(xls_content, csv_content)

        if processed_xls is not None:
            # Lire le fichier Excel traité
            wb = load_workbook(processed_xls)
            output = io.BytesIO()
            wb.save(output)  # Sauvegardez le fichier dans un flux binaire
            wb.close()

            # Réinitialisez le curseur du flux binaire
            output.seek(0)

            # Encodez le contenu en base64 pour le téléchargement
            file_data = output.getvalue()

            # Ajoutez un bouton de téléchargement
            st.success("✅ Félicitations ! Les notes ont été saisies avec succès.")
            file_name = st.text_input("Saisir le nom du fichier Excel de l'adminstration sans l'extension '.xlsx' puis valider.")
            st.download_button(
                label="📥 Télécharger le fichier Excel traité",
                data=file_data,
                file_name=file_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        if nb_none > 0:
            st.warning(f"Attention! {nb_none} étudiants ont été mal identifiés. Vérifiez leurs copies.")

elif section == "Statistiques":
    st.header("📊 Statistiques des notes")
    st.info(
        """
        Télécharger le fichier Excel de l'administration.
        """
    )
    uploaded_excel_file2 = st.file_uploader(
        " ",
        type="xlsx", 
        key="excel_uploader2"
    )
    st.info(
        """
        Télécharger le fichier CSV des notes calculées par AMC.
        """
    )
    uploaded_csv_file = st.file_uploader(
        " ",
        type="csv", 
        key="csv_uploader"
    )
    if uploaded_excel_file2 is not None:
        with st.spinner("Traitement automatique du fichier Excel en cours..."):
            xls, liste = process_excel(uploaded_excel_file2)
    if uploaded_csv_file is not None and uploaded_excel_file2 is not None:
        with st.spinner("Intégration des notes aux étudiants..."):
            csv_clean, anomalies, Notes = process_csv(uploaded_csv_file)
                
            # Générer le fichier Excel final avec en-tête personnalisé
            if Notes is not None:
                # Mettre à jour le fichier Excel avec les notes
                updated_df = update_excel_with_notes(uploaded_excel_file2, Notes)

            st.info(
                """
                Quelques statistiques sur les notes.
                """
            )
            # Affichage des statistiques
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Effectif total", len(xls) if xls is not None else 0)
            with col2:
                st.metric("Présents", len(csv_clean) if csv_clean is not None else 0)
            with col3:
                st.metric("Taux de réussite (%)", round((csv_clean['Note'] >= 10).mean()*100,2)
 if xls is not None and csv_clean is not None else 0)
            with col4:
                st.metric("Mal identifiés", len(anomalies) if anomalies is not None else 0)


            # Calcul des effectifs
            effectifs = csv_clean['Note'].value_counts().reset_index()#.sort_index()
            modalites = csv_clean['Note'].unique()
            effectifs.columns = ['Valeur', 'Effectif']


            # Création du graphique Plotly avec les effectifs affichés sur les barres
            fig = px.bar(effectifs,
                x='Valeur',
                y='Effectif',
                title=" ",
                labels={'Valeur': 'Notes', 'Effectif': 'Effectifs'},
                text_auto=True
            )

            # Personnalisation du layout
            fig.update_layout(
                 title_font_size=20,
                 xaxis_title_font=dict(size=14),
                 yaxis_title_font=dict(size=14),
                 showlegend=False
            )

            # Ajuster la position et le style des étiquettes (optionnel)
            fig.update_traces(textfont_size=14, textangle=0, textposition="outside", width=0.5)
            fig.update_xaxes(tickmode='array', tickvals=list(range(21)), ticktext=[str(i) for i in range(21)])
            fig.update_layout(width=800, height=600)

            st.info(
                """
                Distribution des notes.
                """
            )

            st.plotly_chart(fig)

            # Slider pour ajouter des points
            ajout_points = st.slider("Ajouter des points", min_value=0.0, max_value=5.0, value=0.0, step=0.5)

            if ajout_points > 0:
                csv_plus = csv_clean.copy()

                st.info(
                    """
                    Simulation de la distribution des notes et du taux de réussite après ajout de points.
                    """
                )

                # Ajout des points avec validation (limite maximale de 20)
                csv_plus['Note'] = csv_plus['Note'].apply(lambda x: min(x + ajout_points, 20))

                # Calcul des effectifs après modification
                effectifs_plus = csv_plus['Note'].value_counts().reset_index()
                effectifs_plus.columns = ['Valeur', 'Effectif']

                # Création du graphique Plotly avec les effectifs affichés sur les barres
                fig_plus = px.bar(
                    effectifs_plus,
                    x='Valeur',
                    y='Effectif',
                    title=" ",
                    labels={'Valeur': 'Notes', 'Effectif': 'Effectifs'},
                    text_auto=True
                )

                # Personnalisation du layout
                fig_plus.update_layout(
                    title_font_size=20,
                    xaxis_title_font=dict(size=14),
                    yaxis_title_font=dict(size=14),
                    showlegend=False
                )

                # Ajuster la position et le style des étiquettes
                fig_plus.update_traces(textfont_size=14, textangle=0, textposition="outside", width=0.5)

                # Configuration de l'axe des abscisses pour inclure toutes les valeurs de 0 à 20
                fig_plus.update_xaxes(tickmode='array', tickvals=list(range(21)), ticktext=[str(i) for i in range(21)])

                # Définir la taille du graphique
                fig_plus.update_layout(width=800, height=600)

                # Affichage du taux de réussite mis à jour
                st.metric("Nouveau taux de réussite (%)", round((csv_plus['Note'] >= 10).mean() * 100, 2))

                # Affichage du graphique
                st.plotly_chart(fig_plus)
