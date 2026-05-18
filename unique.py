import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers as xl_numbers
import plotly.express as px
import io
import csv

# =============================================================================
# CONFIGURATION DE LA PAGE
# =============================================================================
st.set_page_config(
    page_title="Outils AMC",
    page_icon="🎓",
    layout="wide"
)

# =============================================================================
# FONCTIONS UTILITAIRES
# =============================================================================

def detect_delimiter(file_content: bytes) -> str:
    """Détecte automatiquement le séparateur du fichier CSV."""
    sample = file_content.decode('utf-8', errors='ignore')
    try:
        dialect = csv.Sniffer().sniff(sample)
        return dialect.delimiter
    except csv.Error:
        return ','


def normalize_code(val) -> str:
    """
    Normalise un code étudiant en chaîne de caractères propre.
    Gère les cas : entier, flottant (123.0 → '123'), texte.
    """
    if val is None:
        return ''
    s = str(val).strip()
    # Supprimer le suffixe '.0' introduit par pandas/Excel pour les entiers lus comme float
    if s.endswith('.0'):
        s = s[:-2]
    return s


# =============================================================================
# RUBRIQUE 1 — LISTE ÉTUDIANTS
# =============================================================================

def process_excel(file) -> tuple:
    """
    Lit le fichier Excel administratif et produit la liste des étudiants
    au format attendu par Auto Multiple Choice.
    Retourne (dataframe_brut, dataframe_liste) ou (None, None) en cas d'erreur.
    """
    try:
        xls = pd.read_excel(file, header=None, dtype=str)  # dtype=str : évite les conversions automatiques

        # Localiser la ligne d'en-tête contenant Code, Nom, Prénom
        header_index = next(
            (idx for idx, row in xls.iterrows()
             if all(col in row.values for col in ['Code', 'Nom', 'Prénom'])),
            None
        )
        if header_index is None:
            st.error("❌ Les colonnes 'Code', 'Nom', 'Prénom' sont introuvables dans le fichier.")
            return None, None

        xls.columns = xls.iloc[header_index]
        xls = xls.iloc[header_index + 1:].reset_index(drop=True)

        if xls.empty:
            st.error("❌ Aucune donnée valide après traitement.")
            return None, None

        missing = [c for c in ['Code', 'Nom', 'Prénom'] if c not in xls.columns]
        if missing:
            st.error(f"❌ Colonnes manquantes : {', '.join(missing)}")
            return None, None

        liste = xls.dropna(subset=['Nom', 'Prénom', 'Code']).copy()
        liste['Code'] = liste['Code'].apply(normalize_code)
        liste['Name'] = liste['Code'] + ' ' + liste['Nom'] + ' ' + liste['Prénom']
        liste = liste[['Code', 'Name']].drop_duplicates()

        return xls, liste

    except Exception as e:
        st.error(f"❌ Erreur lors de la lecture du fichier Excel : {e}")
        return None, None


# =============================================================================
# RUBRIQUE 2 — STATISTIQUES
# =============================================================================

def process_csv(csv_file) -> tuple:
    """
    Lit le fichier CSV d'AMC et retourne un DataFrame propre des notes
    ainsi que les lignes anomalies (Code = NONE).
    """
    try:
        csv_content = csv_file.read()
        delimiter = detect_delimiter(csv_content)
        df = pd.read_csv(io.StringIO(csv_content.decode('utf-8')), delimiter=delimiter)

        if 'Mark' in df.columns:
            df = df.rename(columns={'Mark': 'Note'})

        if 'A:Code' not in df.columns or 'Note' not in df.columns:
            st.error("❌ Colonnes requises 'A:Code' et/ou 'Note' absentes du fichier CSV.")
            return None, None

        anomalies = df[df['A:Code'] == 'NONE'].copy()
        df_clean = df[df['A:Code'] != 'NONE'].copy()

        df_clean['A:Code'] = pd.to_numeric(df_clean['A:Code'], errors='coerce')
        df_clean['Note'] = (
            df_clean['Note'].astype(str)
            .str.replace(',', '.', regex=False)
            .astype(float)
        )

        if df_clean.empty:
            st.error("❌ Aucune donnée valide après nettoyage.")
            return None, None

        return df_clean, anomalies

    except Exception as e:
        st.error(f"❌ Erreur lors de la lecture du fichier CSV : {e}")
        return None, None


def afficher_statistiques(df_notes, anomalies, label=""):
    """Affiche les métriques et le diagramme en bâtons des notes."""
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Présents", len(df_notes))
    with col2:
        valides = (df_notes['Note'] >= 10).sum()
        st.metric("Validés (≥ 10)", int(valides))
    with col3:
        taux = round((valides / len(df_notes)) * 100, 2) if len(df_notes) > 0 else 0
        st.metric("Taux de réussite", f"{taux} %")
    with col4:
        st.metric("Mal identifiés", len(anomalies) if anomalies is not None else 0)

    effectifs = df_notes['Note'].value_counts().reset_index()
    effectifs.columns = ['Note', 'Effectif']

    fig = px.bar(
        effectifs, x='Note', y='Effectif',
        title=f"Distribution des notes{' — ' + label if label else ''}",
        labels={'Note': 'Notes', 'Effectif': 'Effectifs'},
        text_auto=True
    )
    fig.update_layout(
        title_font_size=18,
        xaxis_title_font=dict(size=13),
        yaxis_title_font=dict(size=13),
        showlegend=False,
        width=800, height=500
    )
    fig.update_traces(textfont_size=13, textangle=0, textposition="outside", width=0.5)
    fig.update_xaxes(
        tickmode='array',
        tickvals=list(range(21)),
        ticktext=[str(i) for i in range(21)]
    )
    st.plotly_chart(fig, use_container_width=True)


# =============================================================================
# RUBRIQUE 3 — TRANSFERT DES NOTES
# =============================================================================

def process_csv2excel(xls_file, csv_file, add_notes: float = 0.0) -> tuple:
    """
    Fusionne les notes du CSV AMC vers le fichier Excel administratif.

    Correction clé : les codes sont normalisés en texte (normalize_code)
    des deux côtés avant comparaison, ce qui résout le problème de type
    texte vs numérique introduit par la nouvelle version du fichier Excel.

    Retourne (BytesIO, nb_anomalies, nb_transferts, nb_notes_dispo)
    ou (None, 0, 0, 0) en cas d'erreur.
    """
    try:
        # --- Lecture et nettoyage du CSV ---
        csv_content = csv_file.read()
        delimiter = detect_delimiter(csv_content)
        csv_data = pd.read_csv(io.StringIO(csv_content.decode('utf-8')), delimiter=delimiter)

        if 'Mark' in csv_data.columns:
            csv_data = csv_data.rename(columns={'Mark': 'Note'})

        if 'A:Code' not in csv_data.columns or 'Note' not in csv_data.columns:
            st.error("❌ Le fichier CSV ne contient pas les colonnes 'A:Code' ou 'Note'.")
            return None, 0, 0, 0

        anomalies_count = (csv_data['A:Code'] == 'NONE').sum()
        csv_clean = csv_data[csv_data['A:Code'] != 'NONE'].copy()

        # Construction du dictionnaire {code_normalisé → note_float}
        notes_dict = {}
        for _, row in csv_clean.iterrows():
            code_key = normalize_code(row['A:Code'])
            raw_note = row['Note']

            # Conversion de la note en float
            if isinstance(raw_note, str):
                try:
                    note_val = float(raw_note.replace(',', '.').strip())
                except ValueError:
                    note_val = raw_note  # garder le texte si non numérique
            else:
                note_val = float(raw_note) if raw_note is not None else raw_note

            # Ajout du bonus (plafonné à 20)
            if add_notes > 0 and isinstance(note_val, float):
                note_val = min(note_val + add_notes, 20.0)

            notes_dict[code_key] = note_val

        if not notes_dict:
            st.warning("⚠️ Aucune note valide dans le fichier CSV.")
            return None, 0, 0, 0

        # --- Ouverture du fichier Excel ---
        wb = load_workbook(xls_file)
        ws = wb.active

        # Recherche dynamique des colonnes Code et Note (15 premières lignes)
        code_col_idx = note_col_idx = header_row_idx = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15), start=1):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip().lower()
                    if val == 'code' and code_col_idx is None:
                        code_col_idx = cell.column
                        header_row_idx = row_idx
                    elif val == 'note' and note_col_idx is None:
                        note_col_idx = cell.column
                        header_row_idx = row_idx
            if code_col_idx and note_col_idx:
                break

        if not code_col_idx or not note_col_idx:
            st.error("❌ Colonnes 'Code' et/ou 'Note' introuvables dans le fichier Excel.")
            return None, 0, 0, 0

        # --- Transfert des notes ---
        matched_count = 0
        max_col = max(code_col_idx, note_col_idx)

        for row in ws.iter_rows(min_row=header_row_idx + 1, max_col=max_col):
            code_cell = row[code_col_idx - 1]
            note_cell = row[note_col_idx - 1]

            if code_cell.value is None:
                continue

            # Normalisation du code Excel en texte — correction du problème texte/numérique
            excel_code = normalize_code(code_cell.value)

            if excel_code in notes_dict:
                final_note = notes_dict[excel_code]
                if isinstance(final_note, float) and final_note == int(final_note):
                    note_cell.value = int(final_note)
                    note_cell.number_format = '0'
                else:
                    note_cell.value = final_note
                    note_cell.number_format = '0.00'
                matched_count += 1

        # --- Sauvegarde ---
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        return output, int(anomalies_count), matched_count, len(notes_dict)

    except Exception as e:
        st.error(f"❌ Erreur technique : {e}")
        return None, 0, 0, 0


# =============================================================================
# INTERFACE UTILISATEUR
# =============================================================================

st.title("🎓 Outils AMC — Gestion des notes")

section = st.sidebar.radio(
    "Navigation",
    ["👨‍🎓 Liste étudiants", "📊 Statistiques des notes", "✍️ Transfert des notes"],
    index=0
)

# ---------------------------------------------------------------------------
# RUBRIQUE 1 — LISTE ÉTUDIANTS
# ---------------------------------------------------------------------------
if section == "👨‍🎓 Liste étudiants":
    st.header("👨‍🎓 Génération de la liste étudiants pour AMC")

    st.info(
        "**Objectif :** Convertir le fichier Excel de l'administration en liste CSV "
        "exploitable par Auto Multiple Choice (AMC).\n\n"
        "**Format produit :** deux colonnes — `Code` et `Name` (Code + Nom + Prénom)."
    )

    uploaded_excel = st.file_uploader(
        "📄 Charger le fichier Excel de l'administration (.xlsx)",
        type="xlsx",
        key="excel_etudiants"
    )

    if uploaded_excel:
        with st.spinner("Lecture du fichier en cours…"):
            xls, liste = process_excel(uploaded_excel)

        if xls is not None and liste is not None:
            st.success(f"✅ Fichier lu avec succès — **{len(liste)} étudiants** trouvés.")

            with st.expander("🔎 Aperçu du fichier Excel brut (10 premières lignes)"):
                st.dataframe(xls.head(10), use_container_width=True)

            st.subheader("Liste générée pour AMC")
            st.dataframe(liste.head(10), use_container_width=True)

            csv_bytes = liste.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Télécharger la liste au format CSV",
                data=csv_bytes,
                file_name="liste_etudiants_amc.csv",
                mime="text/csv"
            )

# ---------------------------------------------------------------------------
# RUBRIQUE 2 — STATISTIQUES
# ---------------------------------------------------------------------------
elif section == "📊 Statistiques des notes":
    st.header("📊 Statistiques des notes")

    st.info(
        "**Objectif :** Analyser la distribution des notes issues d'AMC et simuler "
        "l'impact d'un ajout de points.\n\n"
        "**Fichier attendu :** export CSV standard d'AMC (colonnes `A:Code` et `Note` ou `Mark`)."
    )

    uploaded_csv = st.file_uploader(
        "📄 Charger le fichier CSV des notes AMC",
        type="csv",
        key="csv_stats"
    )

    if uploaded_csv:
        with st.spinner("Analyse en cours…"):
            df_notes, anomalies = process_csv(uploaded_csv)

        if df_notes is not None:
            st.success(f"✅ Fichier lu avec succès — **{len(df_notes)} étudiants présents**.")

            st.subheader("Distribution des notes (résultats bruts)")
            afficher_statistiques(df_notes, anomalies)

            st.divider()
            st.subheader("Simulation — Ajout de points")
            ajout = st.slider(
                "Points à ajouter à chaque étudiant (plafond : 20/20)",
                min_value=0.0, max_value=5.0, value=0.0, step=0.5
            )

            if ajout > 0:
                df_sim = df_notes.copy()
                df_sim['Note'] = df_sim['Note'].apply(lambda x: min(x + ajout, 20))
                st.subheader(f"Distribution simulée après +{ajout} point(s)")
                afficher_statistiques(df_sim, anomalies, label=f"+{ajout} pt(s)")

# ---------------------------------------------------------------------------
# RUBRIQUE 3 — TRANSFERT DES NOTES
# ---------------------------------------------------------------------------
elif section == "✍️ Transfert des notes":
    st.header("✍️ Transfert des notes vers le fichier Excel")

    st.info(
        "**Objectif :** Reporter automatiquement les notes calculées par AMC dans le fichier "
        "Excel fourni par l'administration.\n\n"
        "**Fichiers attendus :**\n"
        "- Excel administration : doit contenir les colonnes `Code` et `Note`.\n"
        "- CSV AMC : export standard avec colonnes `A:Code` et `Note` (ou `Mark`).\n\n"
        "⚠️ Si certains étudiants sont signalés *mal identifiés*, leurs notes devront être "
        "saisies manuellement."
    )

    col_left, col_right = st.columns(2)
    with col_left:
        xls_file = st.file_uploader(
            "📄 Fichier Excel de l'administration (.xlsx)",
            type=["xlsx", "xls"],
            key="xls_notes"
        )
    with col_right:
        csv_file = st.file_uploader(
            "📄 Fichier CSV des notes AMC (.csv)",
            type="csv",
            key="csv_notes"
        )

    add_notes = st.number_input(
        "➕ Points bonus à ajouter (0 = aucun, maximum 5)",
        min_value=0.0, max_value=5.0, value=0.0, step=0.5
    )

    btn_disabled = not (xls_file and csv_file)
    if st.button("🚀 Lancer le transfert", type="primary", disabled=btn_disabled):
        xls_file.seek(0)
        csv_file.seek(0)

        with st.spinner("Transfert en cours…"):
            result, nb_anomalies, nb_transferts, nb_dispo = process_csv2excel(
                xls_file, csv_file, add_notes
            )

        if result:
            st.success(
                f"✅ Transfert réussi — **{nb_transferts} notes** insérées "
                f"sur {nb_dispo} disponibles dans le CSV."
            )
            if nb_anomalies > 0:
                st.warning(
                    f"⚠️ **{nb_anomalies} étudiant(s) mal identifié(s)** (code = NONE). "
                    "Vérifiez leurs copies et saisissez leurs notes manuellement."
                )

            nom_fichier = st.text_input(
                "💾 Nom du fichier de sortie (sans extension)",
                value="notes_finales"
            )
            st.download_button(
                label="📥 Télécharger le fichier Excel avec les notes",
                data=result,
                file_name=f"{nom_fichier}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("❌ Le transfert a échoué. Vérifiez les fichiers chargés.")
