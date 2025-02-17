import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
import io


# Fonction de traitement pour le fichier Excel
def process_excel(file):
    try:
        # Lire le fichier Excel sans header
        xls = pd.read_excel(file, header=None)

        # Trouver l'index de la ligne d'en-tête
        header_index = next(
            (idx for idx, row in xls.iterrows() if all(col in row.values for col in ['Code', 'Nom', 'Prénom'])),
            None
        )

        if header_index is None:
            st.error("Les colonnes 'Code', 'Nom', 'Prénom' sont introuvables dans le fichier.")
            return None, None

        # Redéfinir les en-têtes et supprimer les lignes précédentes
        xls.columns = xls.iloc[header_index]
        xls = xls.iloc[header_index + 1:].reset_index(drop=True)

        # Vérification si le fichier est vide après nettoyage
        if xls.empty:
            st.error("Aucune donnée valide après le traitement des lignes.")
            return None, None

        # Vérification des colonnes nécessaires (double vérification)
        required_columns = ['Nom', 'Prénom', 'Code']
        missing = [col for col in required_columns if col not in xls.columns]
        if missing:
            st.error(f"Colonnes manquantes après traitement : {', '.join(missing)}")
            return None, None

        # Nettoyage des données
        liste = xls.dropna(subset=['Nom', 'Prénom', 'Code'])
        liste['Name'] = liste['Code'].astype(str) + ' ' + liste['Nom'] + ' ' + liste['Prénom']
        liste = liste[['Code', 'Name']].drop_duplicates()
        return xls, liste

    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Excel : {str(e)}")
        st.info("Assurez-vous que le fichier est bien formaté et contient les colonnes requises.")
        return None, None


# Fonction de traitement pour le fichier CSV
def process_csv2excel(xls_file, csv_file, add_notes=0):
    try:
        # Lire le fichier CSV avec A:Code comme float
        csv = pd.read_csv(csv_file, delimiter=';', encoding='utf-8', dtype={'A:Code': float})

        # Nettoyer les données : supprimer les lignes où 'A:Code' == 'NONE'
        anomalies = csv[pd.isna(csv['A:Code']) | (csv['A:Code'] == 0)].copy()  # Traiter les NaN ou les 0 comme anomalies
        csv_clean = csv[pd.notna(csv['A:Code']) & (csv['A:Code'] != 0)].copy()

        if add_notes > 0:
            # Ajout des points avec validation (limite maximale de 20)
            csv_clean['Note'] = csv_clean['Note'].apply(lambda x: min(x + add_notes, 20) if pd.notna(x) else x)

        # Vérifier si le fichier nettoyé est vide
        if csv_clean.empty:
            st.error("Aucune donnée valide après le nettoyage !")
            return None, None

        # Construire le dictionnaire Notes
        notes_etudiants = {
            row['A:Code']: row['Note']
            for _, row in csv_clean.iterrows()
            if pd.notna(row['A:Code'])
        }

        # Charger le fichier Excel avec Code comme float
        xls = pd.read_excel(xls_file, dtype={'Code': float})
        wb = load_workbook(filename=xls_file)
        sheet = wb.active

        # Trouver la colonne "Note"
        cell_row, cell_column = None, None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 'Note':  # Vérifier si la valeur de la cellule est 'Note'
                    cell_row = cell.row
                    cell_column = cell.column
                    break
            if cell_row is not None and cell_column is not None:
                break

        if cell_row is None or cell_column is None:
            st.warning("L'en-tête 'Note' n'a pas été trouvé dans le fichier.")
            return None, None

        # Parcourir les lignes du fichier Excel
        for row in sheet.iter_rows(min_row=cell_row + 1, values_only=False):
            code_etudiant_cell = row[0]  # Colonne des codes d'étudiants
            note_cell = row[cell_column - 1]  # Colonne des notes

            # Vérifier si le code étudiant existe dans le dictionnaire
            if pd.notna(code_etudiant_cell.value) and code_etudiant_cell.value in notes_etudiants:
                note_cell.value = notes_etudiants[code_etudiant_cell.value]

        # Enregistrer les modifications dans le fichier Excel
        output = io.BytesIO()
        wb.save(output)
        wb.close()

        output.seek(0)
        return output, len(anomalies)

    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Excel : {str(e)}")
        st.error("Assurez-vous que le fichier est bien formaté et contient les colonnes requises.")
        return None, None
    
# Fonction de traitement pour le fichier CSV
def process_csv(csv_file):
    try:
        csv = pd.read_csv(csv_file, delimiter=';', encoding='utf-8')

        # Nettoyer les données : supprimer les lignes où 'A:Code' == 'NONE'
        csv_nones = csv[csv['A:Code'] == 'NONE'].copy()
        csv_clean = csv[csv['A:Code'] != 'NONE'].copy()

        # Vérifier si le fichier nettoyé est vide
        if csv_clean.empty:
            st.error("Aucune donnée valide après le nettoyage !")

        return csv_clean, csv_nones

    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Excel : {str(e)}")
        st.error("Assurez-vous que le fichier est bien formaté et contient les colonnes requises.")
        return None, None


# ----------------- Interface utilisateur -----------------
st.title("🛠️ Outils pour AMC")

# Sidebar pour les sections
section = st.sidebar.radio(" ", ["ETUDIANTS", "STATISTIQUES", "NOTES"])

if section == "ETUDIANTS":
    st.header("👨‍🎓 Liste des étudiants")
    st.info(
        """
        Pour générer la liste des étudiants au format CSV à charger dans Auto Multiple Choice, télécharger le fichier Excel envoyé par l'administration.
        """
    )

    uploaded_excel_file = st.file_uploader(
        " ",
        type="xlsx",
        key="excel_uploader"
    )

    if uploaded_excel_file is not None:
        with st.spinner("Traitement automatique du fichier Excel en cours..."):
            xls, liste = process_excel(uploaded_excel_file)

            if xls is not None:
                st.success(f"✅  Lecture du fichier Excel réussie !")
                st.write("🔎 Aperçu de la base de données des étudiants:")
                st.write(xls.head(10))
                st.write("🔎 Aperçu de la liste des étudiants à fournir à Auto Multiple Choice:")
                st.write(liste.head(10))
                st.success(f"🔢 La liste contient {len(xls)} étudiants.")
                # Générer le fichier CSV
                csv_data = liste.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📥 Télécharger la liste des étudiants au format CSV",
                    data=csv_data,
                    file_name="liste_etudiants.csv",
                    mime="text/csv"
                )

elif section == "STATISTIQUES":
    st.header("📊 Statistiques des notes")

    st.info(
        """
        Pour avoir une idée sur la distribution statistique des notes, télécharger le fichier des notes calculées par Auto Multiple Choice au format CSV.
        """
    )
    uploaded_csv_file = st.file_uploader(
        " ",
        type="csv",
        key="csv_uploader"
    )

    if uploaded_csv_file is not None:
        with st.spinner("Intégration des notes aux étudiants..."):
            csv_clean, csv_nones = process_csv(uploaded_csv_file)


            # Calcul des effectifs
            effectifs = csv_clean['Note'].value_counts().reset_index()  # .sort_index()
            modalites = csv_clean['Note'].unique()
            effectifs.columns = ['Valeur', 'Effectif']

            # Création du graphique Plotly avec les effectifs affichés sur les barres
            fig = px.bar(effectifs,
                         x='Valeur',
                         y='Effectif',
                         title=" ",
                         labels={'Valeur': 'Notes', 'Effectif': 'Effectifs'},
                         text_auto=True
                         )

            # Personnalisation du layout
            fig.update_layout(
                title_font_size=20,
                xaxis_title_font=dict(size=14),
                yaxis_title_font=dict(size=14),
                showlegend=False
            )

            # Ajuster la position et le style des étiquettes (optionnel)
            fig.update_traces(textfont_size=14, textangle=0, textposition="outside", width=0.5)
            fig.update_xaxes(tickmode='array', tickvals=list(range(21)), ticktext=[str(i) for i in range(21)])
            fig.update_layout(width=800, height=600)

            st.plotly_chart(fig)

            st.info(
                """
                Quelques statistiques des notes.
                """
            )

            # Affichage des statistiques
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Présents", len(csv_clean) if csv_clean is not None else 0)
            with col2:
                st.metric("Validés", (csv_clean['Note'] >= 10).sum() if csv_clean is not None else 0)
            with col3:
                st.metric("Taux de réussite (%)", round(((csv_clean['Note'] >= 10).sum() / len(csv_clean)) * 100,
                                                        2) if csv_clean is not None else 0)
            with col4:
                st.metric("Mal identifiés", len(csv_nones) if csv_nones is not None else 0)


            # Slider pour ajouter des points
            ajout_points = st.slider("Ajouter des points", min_value=0.0, max_value=5.0, value=0.0, step=0.5)

            if ajout_points > 0:
                csv_plus = csv_clean.copy()

                st.info(
                    """
                    Simulation de la distribution des notes et du taux de réussite après ajout de points.
                    """
                )

                # Ajout des points avec validation (limite maximale de 20)
                csv_plus['Note'] = csv_plus['Note'].apply(lambda x: min(x + ajout_points, 20))

                # Calcul des effectifs après modification
                effectifs_plus = csv_plus['Note'].value_counts().reset_index()
                effectifs_plus.columns = ['Valeur', 'Effectif']

                # Création du graphique Plotly avec les effectifs affichés sur les barres
                fig_plus = px.bar(
                    effectifs_plus,
                    x='Valeur',
                    y='Effectif',
                    title=" ",
                    labels={'Valeur': 'Notes', 'Effectif': 'Effectifs'},
                    text_auto=True
                )

                # Personnalisation du layout
                fig_plus.update_layout(
                    title_font_size=20,
                    xaxis_title_font=dict(size=14),
                    yaxis_title_font=dict(size=14),
                    showlegend=False
                )

                # Ajuster la position et le style des étiquettes
                fig_plus.update_traces(textfont_size=14, textangle=0, textposition="outside", width=0.5)

                # Configuration de l'axe des abscisses pour inclure toutes les valeurs de 0 à 20
                fig_plus.update_xaxes(tickmode='array', tickvals=list(range(21)), ticktext=[str(i) for i in range(21)])

                # Définir la taille du graphique
                fig_plus.update_layout(width=800, height=600)

                # Affichage du taux de réussite mis à jour
                st.metric("Nouveau taux de réussite (%)", round((csv_plus['Note'] >= 10).mean() * 100, 2))

                # Affichage du graphique
                st.plotly_chart(fig_plus)




elif section == "NOTES":
    st.header("✍️ Traitement des notes")
    st.info(
        """
        Télécharger le fichier Excel envoyé par l'administration pour la saisie des notes.
        """
    )
    xls_file = st.file_uploader(
        " ",
        type="xlsx",
        key="excel_uploader2"
    )

    st.info(
        """
        Télécharger le fichier des notes calculées par Auto Multiple Choice au format CSV.
        """
    )
    csv_file = st.file_uploader(
        "",
        type="csv",
        key="csv_uploader"
    )
    if xls_file is not None and csv_file is not None:
        with st.spinner("Traitement automatique du fichier Excel en cours..."):
            # Lire le contenu du fichier CSV en mémoire
            csv_content = io.BytesIO(csv_file.getvalue())

            # Lire le contenu du fichier Excel en mémoire
            xls_content = io.BytesIO(xls_file.getvalue())

            # Traitez les fichiers avec process_csv2excel
            add_notes = st.number_input("Combien voulez-vous ajouter de points à l'ensemble des étudiants?", step=0.5)
            processed_xls, nb_none = process_csv2excel(xls_content, csv_content, add_notes)

        if processed_xls is not None:
            # Lire le fichier Excel traité
            wb = load_workbook(processed_xls)
            output = io.BytesIO()
            wb.save(output)  # Sauvegardez le fichier dans un flux binaire
            wb.close()

            # Réinitialisez le curseur du flux binaire
            output.seek(0)

            # Encodez le contenu en base64 pour le téléchargement
            file_data = output.getvalue()

            # Ajoutez un bouton de téléchargement
            st.success("✅ Félicitations ! Les notes ont été saisies avec succès.")
            file_name = st.text_input(
                "Saisir le nom du fichier Excel de l'adminstration sans l'extension '.xlsx' puis valider.")
            st.download_button(
                label="📥 Télécharger le fichier Excel traité",
                data=file_data,
                file_name=file_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        if nb_none > 0:
            st.warning(f"Attention! {nb_none} étudiants ont été mal identifiés. Vérifiez leurs copies et saisissez leurs notes manuellement.")