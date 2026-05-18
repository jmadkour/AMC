"""
Outils AMC — Gestion des notes
================================
Refactoring complet :
  - @st.cache_data sur tous les traitements fichiers
  - Détection d'encodage (chardet) et de délimiteur robuste
  - normalize_code : gestion NaN, zéros de tête, types mixtes
  - Rapport de réconciliation avant toute écriture Excel
  - Statistiques enrichies (médiane, σ, percentiles, seuil visuel)
  - Export statistiques en Excel
  - Sélection de feuille Excel
  - transfer_to_excel isolé et testable indépendamment
"""

from __future__ import annotations

import csv
import io
from typing import Optional

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook

# chardet est optionnel — fallback sur utf-8 si absent
try:
    import chardet
    _HAS_CHARDET = True
except ImportError:
    _HAS_CHARDET = False

# =============================================================================
# CONFIGURATION
# =============================================================================
st.set_page_config(
    page_title="Outils AMC",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# UTILITAIRES CORE
# =============================================================================

def detect_encoding(content: bytes) -> str:
    """Détecte l'encodage du fichier. Retourne 'utf-8' si chardet est absent."""
    if _HAS_CHARDET:
        result = chardet.detect(content[:8192])
        enc = result.get("encoding") or "utf-8"
        # chardet peut retourner 'ascii' pour du utf-8 pur — on normalise
        return enc if enc.lower() != "ascii" else "utf-8"
    return "utf-8"


def detect_delimiter(content: bytes, encoding: str = "utf-8") -> str:
    """
    Détecte le séparateur CSV.
    Stratégie : csv.Sniffer → comptage de candidats → défaut ','.
    """
    sample = content.decode(encoding, errors="replace")[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        for delim in (";", ",", "\t", "|"):
            if sample.count(delim) >= 2:
                return delim
        return ","


def normalize_code(val) -> str:
    """
    Normalise un code étudiant en str propre :
      - None / NaN          → ''
      - 12345.0 (float)     → '12345'
      - '  012 '            → '012'  (zéros de tête préservés)
      - 12345 (int)         → '12345'
    """
    if val is None:
        return ""
    # NaN float
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    # Supprimer le '.0' uniquement si la partie entière est purement numérique
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Sérialise un DataFrame en bytes .xlsx."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return buf.getvalue()


def find_header_row(
    xls: pd.DataFrame, required_cols: set[str]
) -> Optional[int]:
    """Localise l'index de la ligne contenant les colonnes requises."""
    for idx, row in xls.iterrows():
        if required_cols.issubset(set(row.dropna().astype(str).values)):
            return idx
    return None


def load_excel_with_sheet(
    file_bytes: bytes,
) -> tuple[Optional[pd.DataFrame], list[str], str]:
    """
    Charge le fichier Excel en mode dtype=str.
    Retourne (df, sheet_names, active_sheet_name).
    """
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        sheets = xl.sheet_names
        # On préfère la première feuille non vide
        for sheet in sheets:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None, dtype=str)
            if not df.empty:
                return df, sheets, sheet
        return None, sheets, ""
    except Exception as e:
        return None, [], str(e)


# =============================================================================
# RUBRIQUE 1 — LISTE ÉTUDIANTS
# =============================================================================

@st.cache_data(show_spinner=False)
def process_excel(
    file_bytes: bytes, sheet_name: Optional[str] = None
) -> tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], str]:
    """
    Lit le fichier Excel administratif et produit la liste AMC.

    Retourne (df_brut, df_liste, message).
    df_liste colonnes : Code, Name (Code + Nom + Prénom).
    """
    try:
        xls = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name or 0,
            header=None,
            dtype=str,
        )
    except Exception as e:
        return None, None, f"Impossible de lire le fichier Excel : {e}"

    required = {"Code", "Nom", "Prénom"}
    header_idx = find_header_row(xls, required)

    if header_idx is None:
        return None, None, (
            "En-tête introuvable. "
            "Le fichier doit contenir les colonnes 'Code', 'Nom', 'Prénom'."
        )

    xls.columns = xls.iloc[header_idx].astype(str)
    xls = xls.iloc[header_idx + 1 :].reset_index(drop=True)
    xls.columns.name = None

    missing = required - set(xls.columns)
    if missing:
        return None, None, f"Colonnes manquantes après détection d'en-tête : {missing}"

    liste = xls.dropna(subset=["Code", "Nom", "Prénom"]).copy()
    liste["Code"] = liste["Code"].apply(normalize_code)
    liste = liste[liste["Code"] != ""]

    dupes_before = len(liste)
    liste = liste.drop_duplicates(subset="Code")
    dupes_dropped = dupes_before - len(liste)

    liste["Name"] = (
        liste["Code"]
        + " "
        + liste["Nom"].str.strip()
        + " "
        + liste["Prénom"].str.strip()
    )
    liste = liste[["Code", "Name"]].reset_index(drop=True)

    if liste.empty:
        return None, None, "Aucun étudiant valide après nettoyage."

    msg = f"{len(liste)} étudiants"
    if dupes_dropped:
        msg += f" ({dupes_dropped} doublon(s) de code supprimé(s))"
    return xls, liste, msg


# =============================================================================
# RUBRIQUE 2 — STATISTIQUES
# =============================================================================

@st.cache_data(show_spinner=False)
def process_csv(
    file_bytes: bytes,
) -> tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], str]:
    """
    Lit le fichier CSV AMC.

    Retourne (df_clean, df_anomalies, message).
    df_clean  : étudiants identifiés, colonne 'Note' float.
    df_anomalies : lignes code = NONE.
    """
    try:
        encoding = detect_encoding(file_bytes)
        delimiter = detect_delimiter(file_bytes, encoding)
        df = pd.read_csv(
            io.StringIO(file_bytes.decode(encoding, errors="replace")),
            delimiter=delimiter,
        )
    except Exception as e:
        return None, None, f"Lecture CSV impossible : {e}"

    # Normalisation du nom de la colonne note
    if "Mark" in df.columns and "Note" not in df.columns:
        df = df.rename(columns={"Mark": "Note"})

    missing = [c for c in ("A:Code", "Note") if c not in df.columns]
    if missing:
        return None, None, f"Colonnes requises absentes : {missing}"

    df["A:Code"] = df["A:Code"].astype(str).str.strip()
    mask_none = df["A:Code"].str.upper() == "NONE"
    anomalies = df[mask_none].copy()
    df_clean = df[~mask_none].copy()

    df_clean["Note"] = pd.to_numeric(
        df_clean["Note"].astype(str).str.replace(",", ".", regex=False).str.strip(),
        errors="coerce",
    )
    nb_invalid = df_clean["Note"].isna().sum()
    df_clean = df_clean.dropna(subset=["Note"])

    if df_clean.empty:
        return None, anomalies, "Aucune note valide après nettoyage."

    parts = [f"{len(df_clean)} notes valides"]
    if nb_invalid:
        parts.append(f"{nb_invalid} ligne(s) ignorée(s) (note non numérique)")
    if len(anomalies):
        parts.append(f"{len(anomalies)} mal identifié(s) (code NONE)")
    return df_clean, anomalies, " · ".join(parts)


def compute_stats(notes: pd.Series, seuil: float = 10.0) -> pd.DataFrame:
    """Retourne un DataFrame de statistiques descriptives."""
    valides = (notes >= seuil).sum()
    return pd.DataFrame(
        {
            "Métrique": [
                "Effectif",
                f"Validés (≥ {seuil})",
                "Taux de réussite",
                "Moyenne",
                "Médiane",
                "Écart-type",
                "P25",
                "P75",
                "Min",
                "Max",
            ],
            "Valeur": [
                len(notes),
                int(valides),
                f"{valides / len(notes) * 100:.2f} %",
                f"{notes.mean():.2f}",
                f"{notes.median():.2f}",
                f"{notes.std():.2f}",
                f"{notes.quantile(0.25):.2f}",
                f"{notes.quantile(0.75):.2f}",
                f"{notes.min():.2f}",
                f"{notes.max():.2f}",
            ],
        }
    )


def afficher_statistiques(
    df_notes: pd.DataFrame,
    anomalies: Optional[pd.DataFrame],
    ajout: float = 0.0,
    seuil: float = 10.0,
    label: str = "",
) -> None:
    """Affiche métriques, histogramme et tableau de stats."""
    notes = df_notes["Note"].copy()
    if ajout > 0:
        notes = notes.clip(upper=20.0).add(ajout).clip(upper=20.0)

    valides = (notes >= seuil).sum()
    taux = valides / len(notes) * 100 if len(notes) else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Présents", len(notes))
    c2.metric(f"Validés (≥ {seuil})", int(valides))
    c3.metric("Réussite", f"{taux:.1f} %")
    c4.metric("Moyenne", f"{notes.mean():.2f}")
    c5.metric("Mal identifiés", len(anomalies) if anomalies is not None else 0)

    effectifs = (
        notes.value_counts()
        .reset_index()
        .rename(columns={"index": "Note", "count": "Effectif"})
        .sort_values("Note")
    )
    # Compatibilité pandas < 2.0 et >= 2.0
    if "Note" not in effectifs.columns and notes.name in effectifs.columns:
        effectifs = effectifs.rename(columns={notes.name: "Note"})
    if effectifs.columns.tolist()[:2] == [0, 1]:
        effectifs.columns = ["Note", "Effectif"]

    fig = px.bar(
        effectifs,
        x="Note",
        y="Effectif",
        title=f"Distribution des notes{' — ' + label if label else ''}",
        text_auto=True,
        color="Note",
        color_continuous_scale="Blues",
    )
    fig.add_vline(
        x=seuil,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Seuil {seuil}/20",
        annotation_position="top right",
    )
    fig.update_layout(
        showlegend=False,
        coloraxis_showscale=False,
        height=450,
        xaxis=dict(tickmode="linear", tick0=0, dtick=1, range=[-0.5, 20.5]),
        title_font_size=16,
    )
    fig.update_traces(textfont_size=11, textangle=0, textposition="outside", width=0.6)
    st.plotly_chart(fig, use_container_width=True)

    col_stats, col_anom = st.columns([2, 1])
    with col_stats:
        with st.expander("📋 Statistiques détaillées"):
            st.dataframe(
                compute_stats(notes, seuil),
                use_container_width=True,
                hide_index=True,
            )
    with col_anom:
        if anomalies is not None and len(anomalies) > 0:
            with st.expander(f"⚠️ Mal identifiés ({len(anomalies)})"):
                st.dataframe(anomalies, use_container_width=True, hide_index=True)


# =============================================================================
# RUBRIQUE 3 — TRANSFERT DES NOTES
# =============================================================================

@st.cache_data(show_spinner=False)
def build_notes_dict(
    csv_bytes: bytes, add_notes: float = 0.0
) -> tuple[dict[str, float], int]:
    """
    Construit {code_normalisé: note_float} depuis le CSV AMC.

    Retourne (notes_dict, nb_anomalies).
    Séparé de l'écriture Excel pour faciliter les tests et la mise en cache.
    """
    encoding = detect_encoding(csv_bytes)
    delimiter = detect_delimiter(csv_bytes, encoding)
    df = pd.read_csv(
        io.StringIO(csv_bytes.decode(encoding, errors="replace")),
        delimiter=delimiter,
    )

    if "Mark" in df.columns and "Note" not in df.columns:
        df = df.rename(columns={"Mark": "Note"})

    df["A:Code"] = df["A:Code"].astype(str).str.strip()
    nb_anomalies = int((df["A:Code"].str.upper() == "NONE").sum())
    df = df[df["A:Code"].str.upper() != "NONE"].copy()

    df["_code"] = df["A:Code"].apply(normalize_code)
    df["_note"] = pd.to_numeric(
        df["Note"].astype(str).str.replace(",", ".", regex=False).str.strip(),
        errors="coerce",
    )
    df = df.dropna(subset=["_note"])

    if add_notes > 0:
        df["_note"] = (df["_note"] + add_notes).clip(upper=20.0)

    # En cas de codes dupliqués dans le CSV, on garde la première occurrence
    df = df.drop_duplicates(subset="_code")
    return dict(zip(df["_code"], df["_note"])), nb_anomalies


@st.cache_data(show_spinner=False)
def reconciliation_report(
    xls_bytes: bytes, notes_dict_items: tuple, sheet_name: Optional[str] = None
) -> pd.DataFrame:
    """
    Produit un rapport de correspondance avant toute écriture.

    notes_dict_items est un tuple de (code, note) pour compatibilité cache Streamlit.
    Retourne un DataFrame [Code, Nom, Prénom, Note AMC, Statut].
    """
    notes_dict = dict(notes_dict_items)

    xls = pd.read_excel(
        io.BytesIO(xls_bytes),
        sheet_name=sheet_name or 0,
        header=None,
        dtype=str,
    )
    header_idx = find_header_row(xls, {"Code", "Nom", "Prénom"})
    if header_idx is None:
        return pd.DataFrame()

    xls.columns = xls.iloc[header_idx].astype(str)
    xls = xls.iloc[header_idx + 1 :].reset_index(drop=True)
    xls.columns.name = None

    cols = [c for c in ("Code", "Nom", "Prénom") if c in xls.columns]
    df = xls[cols].copy()
    df["Code"] = df["Code"].apply(normalize_code)
    df = df[df["Code"] != ""].dropna(subset=["Nom", "Prénom"])

    df["Note AMC"] = df["Code"].map(notes_dict)
    df["Statut"] = df["Note AMC"].apply(
        lambda x: "✅ Trouvée" if pd.notna(x) else "❌ Absente"
    )
    df["Note AMC"] = df["Note AMC"].apply(
        lambda x: f"{x:.2f}" if pd.notna(x) else "—"
    )
    return df[["Code", "Nom", "Prénom", "Note AMC", "Statut"]].reset_index(drop=True)


def transfer_to_excel(
    xls_bytes: bytes,
    notes_dict: dict[str, float],
    sheet_name: Optional[str] = None,
) -> tuple[bytes, int]:
    """
    Écrit les notes dans le fichier Excel via openpyxl.

    Retourne (excel_bytes, nb_matches).
    Lève ValueError si les colonnes Code/Note sont introuvables.
    """
    wb = load_workbook(io.BytesIO(xls_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    code_col = note_col = header_row = None

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
        for cell in row:
            if cell.value is not None:
                v = str(cell.value).strip().lower()
                if v == "code" and code_col is None:
                    code_col = cell.column
                    header_row = r_idx
                elif v == "note" and note_col is None:
                    note_col = cell.column
        if code_col and note_col:
            break

    if not (code_col and note_col):
        wb.close()
        raise ValueError(
            "Colonnes 'Code' et/ou 'Note' introuvables dans le fichier Excel. "
            "Vérifiez la feuille sélectionnée."
        )

    matched = 0
    max_col = max(code_col, note_col)

    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col):
        code_cell = row[code_col - 1]
        note_cell = row[note_col - 1]

        if code_cell.value is None:
            continue

        key = normalize_code(code_cell.value)
        if key in notes_dict:
            val = notes_dict[key]
            # Stockage entier si val est un entier exact, sinon 2 décimales
            if val == int(val):
                note_cell.value = int(val)
                note_cell.number_format = "0"
            else:
                note_cell.value = round(val, 2)
                note_cell.number_format = "0.00"
            matched += 1

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), matched


# =============================================================================
# INTERFACE
# =============================================================================

st.title("🎓 Outils AMC — Gestion des notes")

section = st.sidebar.radio(
    "Navigation",
    ["👨‍🎓 Liste étudiants", "📊 Statistiques des notes", "✍️ Transfert des notes"],
)

if not _HAS_CHARDET:
    st.sidebar.warning(
        "⚠️ Module `chardet` non installé. "
        "La détection d'encodage est désactivée (UTF-8 supposé). "
        "`pip install chardet` pour l'activer."
    )

# ---------------------------------------------------------------------------
# RUBRIQUE 1 — LISTE ÉTUDIANTS
# ---------------------------------------------------------------------------
if section == "👨‍🎓 Liste étudiants":
    st.header("👨‍🎓 Génération de la liste étudiants pour AMC")

    st.info(
        "Convertit le fichier Excel de l'administration en liste CSV "
        "exploitable par Auto Multiple Choice.\n\n"
        "**Format produit :** colonnes `Code` et `Name` (Code · Nom · Prénom)."
    )

    uploaded = st.file_uploader(
        "📄 Fichier Excel administration (.xlsx)",
        type=["xlsx", "xls"],
        key="xl_etu",
    )

    if uploaded:
        file_bytes = uploaded.read()

        # Sélection de feuille si nécessaire
        _, sheets, default_sheet = load_excel_with_sheet(file_bytes)
        sheet_name = default_sheet
        if len(sheets) > 1:
            sheet_name = st.selectbox(
                "📑 Feuille à utiliser",
                sheets,
                index=sheets.index(default_sheet) if default_sheet in sheets else 0,
            )

        with st.spinner("Lecture…"):
            xls, liste, msg = process_excel(file_bytes, sheet_name)

        if liste is not None:
            st.success(f"✅ {msg}")

            col1, col2 = st.columns([1, 1])
            with col1:
                with st.expander("🔎 Aperçu brut (10 premières lignes)"):
                    st.dataframe(xls.head(10), use_container_width=True)
            with col2:
                st.subheader("Liste générée")
                st.dataframe(liste, use_container_width=True, height=300)

            st.download_button(
                "📥 Télécharger la liste CSV pour AMC",
                data=liste.to_csv(index=False).encode("utf-8"),
                file_name="liste_etudiants_amc.csv",
                mime="text/csv",
            )
        else:
            st.error(f"❌ {msg}")

# ---------------------------------------------------------------------------
# RUBRIQUE 2 — STATISTIQUES
# ---------------------------------------------------------------------------
elif section == "📊 Statistiques des notes":
    st.header("📊 Statistiques des notes")

    st.info(
        "Analyse la distribution des notes AMC et simule l'impact d'un bonus.\n\n"
        "**Fichier attendu :** export CSV AMC (colonnes `A:Code` et `Note` ou `Mark`)."
    )

    uploaded = st.file_uploader(
        "📄 Fichier CSV notes AMC (.csv)", type="csv", key="csv_stat"
    )

    if uploaded:
        file_bytes = uploaded.read()
        with st.spinner("Analyse…"):
            df_notes, anomalies, msg = process_csv(file_bytes)

        if df_notes is not None:
            st.success(f"✅ {msg}")

            seuil = st.number_input(
                "Seuil de validation (par défaut 10/20)",
                min_value=0.0, max_value=20.0, value=10.0, step=0.5,
            )

            st.subheader("Résultats bruts")
            afficher_statistiques(df_notes, anomalies, seuil=seuil)

            st.download_button(
                "📥 Exporter les statistiques (.xlsx)",
                data=df_to_excel_bytes(compute_stats(df_notes["Note"], seuil)),
                file_name="statistiques_notes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.divider()
            st.subheader("Simulation — Ajout de points")
            ajout = st.slider(
                "Points à ajouter (plafond 20/20)", 0.0, 5.0, 0.0, 0.5
            )

            if ajout > 0:
                st.subheader(f"Distribution simulée — +{ajout} pt(s)")
                afficher_statistiques(
                    df_notes, anomalies, ajout=ajout, seuil=seuil, label=f"+{ajout} pt(s)"
                )
        else:
            st.error(f"❌ {msg}")

# ---------------------------------------------------------------------------
# RUBRIQUE 3 — TRANSFERT DES NOTES
# ---------------------------------------------------------------------------
elif section == "✍️ Transfert des notes":
    st.header("✍️ Transfert des notes vers le fichier Excel")

    st.info(
        "Reporte les notes AMC dans le fichier Excel de l'administration.\n\n"
        "**Fichiers requis :**\n"
        "- Excel admin : colonnes `Code` et `Note`\n"
        "- CSV AMC : colonnes `A:Code` et `Note` (ou `Mark`)\n\n"
        "⚠️ Les étudiants sans identifiant (code NONE) devront être saisis manuellement."
    )

    col_l, col_r = st.columns(2)
    with col_l:
        xls_file = st.file_uploader(
            "📄 Excel administration (.xlsx / .xls)",
            type=["xlsx", "xls"],
            key="xls_tr",
        )
    with col_r:
        csv_file = st.file_uploader(
            "📄 CSV notes AMC (.csv)", type="csv", key="csv_tr"
        )

    add_notes = st.number_input(
        "➕ Points bonus à ajouter à toutes les notes (0–5, plafonné à 20/20)",
        min_value=0.0, max_value=5.0, value=0.0, step=0.5,
    )

    if xls_file and csv_file:
        xls_bytes = xls_file.read()
        csv_bytes = csv_file.read()

        # Sélection de feuille
        _, sheets, default_sheet = load_excel_with_sheet(xls_bytes)
        sheet_name = default_sheet
        if len(sheets) > 1:
            sheet_name = st.selectbox(
                "📑 Feuille Excel à compléter",
                sheets,
                index=sheets.index(default_sheet) if default_sheet in sheets else 0,
                key="sheet_tr",
            )

        # Construction du dictionnaire de notes (mis en cache)
        try:
            notes_dict, nb_anomalies = build_notes_dict(csv_bytes, add_notes)
        except Exception as e:
            st.error(f"❌ Erreur lors de la lecture du CSV : {e}")
            st.stop()

        if not notes_dict:
            st.warning("⚠️ Aucune note valide trouvée dans le fichier CSV.")
            st.stop()

        # ----------------------------------------------------------------
        # RAPPORT DE RÉCONCILIATION (avant toute écriture)
        # ----------------------------------------------------------------
        st.subheader("📋 Rapport de réconciliation")

        recon = reconciliation_report(
            xls_bytes, tuple(notes_dict.items()), sheet_name
        )

        if recon.empty:
            st.error(
                "❌ Impossible de produire le rapport : "
                "colonnes 'Code', 'Nom', 'Prénom' introuvables dans l'Excel."
            )
            st.stop()

        n_found = int((recon["Statut"] == "✅ Trouvée").sum())
        n_missing = int((recon["Statut"] == "❌ Absente").sum())

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Notes dans le CSV", len(notes_dict))
        c2.metric("Étudiants dans l'Excel", len(recon))
        c3.metric("Correspondances", n_found, delta=None)
        c4.metric("Sans correspondance", n_missing, delta=None)

        if nb_anomalies:
            st.warning(
                f"⚠️ {nb_anomalies} copie(s) mal identifiée(s) dans le CSV (code NONE) — "
                "non incluses dans le transfert."
            )

        tab_ok, tab_ko, tab_all = st.tabs(
            [f"✅ Trouvées ({n_found})", f"❌ Absentes ({n_missing})", "📄 Tout (triable)"]
        )
        with tab_ok:
            df_ok = recon[recon["Statut"] == "✅ Trouvée"]
            if df_ok.empty:
                st.info("Aucune correspondance.")
            else:
                st.dataframe(df_ok, use_container_width=True, hide_index=True)
        with tab_ko:
            df_ko = recon[recon["Statut"] == "❌ Absente"]
            if df_ko.empty:
                st.success("Tous les étudiants ont une note.")
            else:
                st.dataframe(df_ko, use_container_width=True, hide_index=True)
        with tab_all:
            st.dataframe(recon, use_container_width=True, hide_index=True)

        st.download_button(
            "📥 Exporter le rapport de réconciliation (.xlsx)",
            data=df_to_excel_bytes(recon),
            file_name="rapport_reconciliation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()

        # ----------------------------------------------------------------
        # BOUTON DE TRANSFERT (confirmation explicite)
        # ----------------------------------------------------------------
        if n_found == 0:
            st.error(
                "❌ Aucune correspondance trouvée entre le CSV et l'Excel. "
                "Vérifiez le format des codes étudiants des deux fichiers."
            )
        else:
            if st.button(
                f"🚀 Confirmer et transférer {n_found} note(s)",
                type="primary",
            ):
                try:
                    with st.spinner("Écriture en cours…"):
                        output_bytes, matched = transfer_to_excel(
                            xls_bytes, notes_dict, sheet_name
                        )
                    st.success(f"✅ {matched} note(s) insérée(s) avec succès.")

                    nom = st.text_input(
                        "💾 Nom du fichier de sortie (sans extension)",
                        value="notes_finales",
                    )
                    st.download_button(
                        "📥 Télécharger le fichier Excel complété",
                        data=output_bytes,
                        file_name=f"{nom}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                except ValueError as ve:
                    st.error(f"❌ {ve}")
                except Exception as e:
                    st.error(f"❌ Erreur inattendue : {e}")
